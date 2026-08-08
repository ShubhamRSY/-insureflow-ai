from __future__ import annotations

import csv
import io
import re
from typing import Any, Optional

from insureflow.ingestion.base import BaseParser
from insureflow.models.submissions import (
    ExtractedField,
    ScheduleItem,
    ScheduleOfValues,
    UnstructuredSubmission,
)


class ExcelParser(BaseParser):
    def parse(self, raw_data: str, submission_id: str) -> UnstructuredSubmission:
        submission = UnstructuredSubmission(
            submission_id=submission_id,
            source="excel_parser",
            document_type="excel_workbook",
            raw_text=raw_data,
        )

        try:
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(raw_data.encode("latin-1")), data_only=True)
            sheet_names: list[str] = []
            unparsed_sheets: list[str] = []
            sov_versions: dict[str, str] = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                table = self._unmerge_and_extract(ws)
                text = self._table_to_markdown(table, sheet_name)
                submission.raw_text += f"\n\n--- Sheet: {sheet_name} ---\n{text}"
                sheet_names.append(sheet_name)
                extracted = self._extract_fields_from_table(table, sheet_name)
                for key, val in extracted.items():
                    submission.extracted_fields.setdefault(key, []).append(
                        ExtractedField(
                            field_name=key,
                            value=str(val),
                            confidence=0.85,
                            context=f"excel:{sheet_name}",
                        )
                    )
                # Surface template shape so silent reparse or template drift is
                # visible instead of dropping sheets without any SOV signal.
                sov = self._build_sov_from_table(table, sheet_name)
                if sov.items:
                    sov_versions[sheet_name] = sov.template_version
                else:
                    unparsed_sheets.append(sheet_name)

            if sheet_names:
                submission.extracted_fields["excel.sheets"] = [
                    ExtractedField(
                        field_name="excel.sheets",
                        value=", ".join(sheet_names),
                        confidence=1.0,
                        context="excel",
                    )
                ]
            if unparsed_sheets:
                submission.extracted_fields["excel.unparsed_sheets"] = [
                    ExtractedField(
                        field_name="excel.unparsed_sheets",
                        value=", ".join(unparsed_sheets),
                        confidence=0.6,
                        context="excel",
                    )
                ]
            if sov_versions:
                submission.extracted_fields["excel.sov_versions"] = [
                    ExtractedField(
                        field_name="excel.sov_versions",
                        value=", ".join(f"{k}={v}" for k, v in sov_versions.items()),
                        confidence=1.0,
                        context="excel",
                    )
                ]
        except Exception:
            submission.raw_text = raw_data

        return submission

    def parse_structured(self, raw_data: str, submission_id: str) -> list[ScheduleOfValues]:
        try:
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(raw_data.encode("latin-1")), data_only=True)
            return self._extract_sovs(wb)
        except Exception:
            return []

    def parse_csv(self, content: str, submission_id: str) -> UnstructuredSubmission:
        submission = UnstructuredSubmission(
            submission_id=submission_id,
            source="csv_parser",
            document_type="excel_workbook",
            raw_text=content,
        )
        reader = csv.DictReader(io.StringIO(content))
        rows = list(reader)
        if rows:
            headers = list(rows[0].keys())
            md = "| " + " | ".join(headers) + " |\n"
            md += "| " + " | ".join("---" for _ in headers) + " |\n"
            for row in rows:
                md += "| " + " | ".join(row.get(h, "") for h in headers) + " |\n"
            submission.raw_text = md

            for key in headers:
                values = [row.get(key, "") for row in rows if row.get(key, "")]
                if values:
                    submission.extracted_fields[key] = [
                        ExtractedField(
                            field_name=key,
                            value=", ".join(values[:10]),
                            confidence=0.8,
                            context="csv",
                        )
                    ]
        return submission

    def _unmerge_and_extract(self, ws: Any) -> list[list[Optional[str]]]:
        try:
            merged = ws.merged_cells.ranges
        except AttributeError:
            merged = []

        unmerge_map: dict[tuple[int, int], str] = {}
        for mr in merged:
            min_col = mr.min_col
            min_row = mr.min_row
            max_col = mr.max_col
            max_row = mr.max_row
            val = ws.cell(min_row, min_col).value
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    unmerge_map[(r, c)] = str(val) if val is not None else ""

        table: list[list[Optional[str]]] = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            cells: list[Optional[str]] = []
            for cell in row:
                key = (cell.row, cell.column)
                if key in unmerge_map:
                    cells.append(unmerge_map[key])
                else:
                    v = cell.value
                    cells.append(str(v) if v is not None else None)
            if any(c is not None for c in cells):
                table.append(cells)
        return table

    def _table_to_markdown(self, table: list[list[Optional[str]]], sheet_name: str) -> str:
        if not table:
            return ""
        header = "| " + " | ".join(str(c or "") for c in table[0]) + " |"
        sep = "| " + " | ".join("---" for _ in table[0]) + " |"
        rows = []
        for row in table[1:]:
            rows.append("| " + " | ".join(str(c or "") for c in row) + " |")
        return f"### Sheet: {sheet_name}\n{header}\n{sep}\n" + "\n".join(rows)

    def _extract_fields_from_table(self, table: list[list[Optional[str]]], sheet_name: str) -> dict[str, Any]:
        fields: dict[str, Any] = {}
        if not table:
            return fields

        cov_keywords = [
            "coverage",
            "limit",
            "deductible",
            "premium",
            "building",
            "contents",
            "liability",
            "property",
            "inland",
            "crime",
        ]
        loc_keywords = [
            "location",
            "address",
            "city",
            "state",
            "zip",
            "occupancy",
            "construction",
            "year built",
            "sqft",
        ]

        for row in table:
            if not row:
                continue
            label = str(row[0] or "").strip().lower()
            values = [str(c or "").strip() for c in row[1:] if c is not None]

            if any(kw in label for kw in loc_keywords) and values:
                clean_label = label.replace(" ", "_").replace("/", "_")
                fields[clean_label] = values[0]

            if any(kw in label for kw in cov_keywords) and values:
                clean_label = label.replace(" ", "_").replace("/", "_")
                fields[clean_label] = values[0]

        for row in table:
            if not row:
                continue
            for kw in ["total", "sum", "aggregate"]:
                if str(row[0] or "").strip().lower().startswith(kw) and len(row) > 1:
                    val = str(row[-1] or "").strip()
                    val = re.sub(r"[^\d.,]", "", val)
                    if val:
                        fields[f"total_{sheet_name.lower().replace(' ', '_')}"] = val

        return fields

    def _extract_sovs(self, wb: Any) -> list[ScheduleOfValues]:
        sovs: list[ScheduleOfValues] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            table = self._unmerge_and_extract(ws)
            if not table:
                continue
            sov = self._build_sov_from_table(table, sheet_name)
            if sov.items:
                sovs.append(sov)
        return sovs

    def _build_sov_from_table(self, table: list[list[Optional[str]]], sheet_name: str) -> ScheduleOfValues:
        coverage_map = {
            "building": "Property",
            "contents": "Property",
            "general liability": "CGL",
            "property": "Property",
            "equipment": "Inland Marine",
            "crime": "Crime",
        }

        header = [str(h or "").strip().lower() for h in (table[0] if table else [])]
        col_desc = -1
        col_val = -1
        col_limit = -1
        for i, h in enumerate(header):
            if "description" in h or "item" in h:
                col_desc = i
            elif "value" in h or "amount" in h:
                col_val = i
            elif "limit" in h:
                col_limit = i

        items: list[ScheduleItem] = []
        coverage_type = "Property"
        for slug, cov in coverage_map.items():
            if slug in sheet_name.lower():
                coverage_type = cov
                break

        # Without description/value columns the sheet is not an SOV template;
        # parsing with -1 indices would silently read trailing cells.
        if col_desc < 0 or col_val < 0:
            return ScheduleOfValues(
                schedule_type=f"Excel SOV ({sheet_name})",
                coverage_type=coverage_type,
                template_version="",
            )

        for row in table[1:]:
            if not row or not any(c is not None for c in row):
                continue
            desc = str(row[col_desc]).strip() if row[col_desc] else ""
            raw_val = str(row[col_val]).strip() if row[col_val] else ""
            raw_lim = str(row[col_limit]).strip() if col_limit >= 0 and row[col_limit] else ""

            if not desc or not raw_val:
                continue

            val = self._parse_currency(raw_val)
            limit = self._parse_currency(raw_lim) if raw_lim else None

            items.append(
                ScheduleItem(
                    item_number=str(len(items) + 1),
                    description=desc,
                    value=val,
                    limit=limit,
                )
            )

        total = sum(i.value for i in items)
        return ScheduleOfValues(
            schedule_type=f"Excel SOV ({sheet_name})",
            coverage_type=coverage_type,
            items=items,
            total_value=total,
            template_version=self._template_version(header),
        )

    @staticmethod
    def _template_version(header: list[str]) -> str:
        import hashlib

        signature = "|".join(h for h in header if h)
        if not signature:
            return ""
        return hashlib.sha1(signature.encode("utf-8")).hexdigest()[:8]

    @staticmethod
    def _parse_currency(s: str) -> float:
        cleaned = re.sub(r"[^\d.,]", "", s)
        if not cleaned:
            return 0.0
        if cleaned.count(".") > 1 and cleaned.count(",") > 0:
            cleaned = cleaned.replace(".", "")
            cleaned = cleaned.replace(",", ".")
        try:
            return float(cleaned)
        except ValueError:
            return 0.0
